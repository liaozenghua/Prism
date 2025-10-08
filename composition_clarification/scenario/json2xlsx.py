import json
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from scenario.scenario_utils import *

def json_to_excel(json_data, output_file):
    # 创建新工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 定义列顺序和标题
    columns = [
        'domain', 'intent', 'example', 'element', 'result',
        'labelled1', 'labelled2', 'labelled3', 'preceding_element', 'hierarchical_elements'
    ]

    # 添加表头
    ws.append(columns)

    # 设置列宽
    column_widths = {
        'domain': 15,
        'intent': 20,
        'example': 65,
        'element': 50,
        'result': 45,
        'labelled1': 45,
        'labelled2': 45,
        'labelled3': 45,
        'preceding_element': 45,
        'hierarchical_elements': 45
    }

    for i, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = column_widths.get(col, 20)

    # 处理JSON数据
    for item in json_data:
        row = []
        for col in columns:
            # 获取值并替换换行符
            value = item.get(col, "")
            # if isinstance(value, str):
            #     value = value.replace("\n", " ")
            row.append(value)

        # 添加数据行
        ws.append(row)

    # 设置自动换行和垂直居中
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(columns)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    # 保存文件
    wb.save(output_file)
    print(f"Excel文件已成功生成: {output_file}")



# 示例使用
if __name__ == "__main__":


    # 加载JSON数据（替换为实际JSON文件路径）
    with open('/mnt/d/study/code/ReST-SDLG/scenario/data_all.json', 'r', encoding='utf-8') as f:
        json_data = json.load(f)

    # 转换并保存为Excel
    json_to_excel(json_data, '/mnt/d/study/code/ReST-SDLG/scenario/labelled.xlsx')